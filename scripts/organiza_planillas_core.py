from __future__ import annotations

import csv
import os
import re
import shutil
from collections import defaultdict
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Iterable

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill
except ImportError as exc:  # pragma: no cover - handled at runtime in packaged builds
    raise RuntimeError(
        "openpyxl is required to read and write Excel files. Install it with "
        "`pip install openpyxl`."
    ) from exc


@dataclass(frozen=True)
class PlanillaRow:
    row_number: int
    dig_id: str
    dig_anio: str
    dig_expediente: str
    dig_id_tramite: str
    dig_tramite: str


def safe_resolve(path: Path) -> Path:
    return path.expanduser().resolve()


def ensure_under_base(base: Path, candidate: Path) -> bool:
    try:
        base_r = base.resolve()
        cand_r = candidate.resolve()
        return cand_r == base_r or base_r in cand_r.parents
    except OSError:
        return False


def load_dotenv(path: Path) -> None:
    if not path.exists():
        return
    try:
        for raw in path.read_text(encoding="utf-8").splitlines():
            line = raw.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            key, value = line.split("=", 1)
            key = key.strip()
            value = value.strip().strip('"').strip("'")
            if key and key not in os.environ:
                os.environ[key] = value
    except OSError:
        return


def _strip_outer_quotes(value: str) -> str:
    return value.strip().strip('"').strip("'")


def _normalize_header(value: object) -> str:
    text = "" if value is None else str(value)
    text = _strip_outer_quotes(text)
    text = text.strip().lower()
    text = re.sub(r"[^a-z0-9_]+", "", text)
    return text


def normalize_cell(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        text = format(value, "f").rstrip("0").rstrip(".")
        return text

    text = _strip_outer_quotes(str(value)).strip()
    if not text:
        return ""

    compact = re.sub(r"\s+", " ", text)
    if re.fullmatch(r"[0-9][0-9.,\s]*", compact):
        return re.sub(r"\D", "", compact)
    if re.fullmatch(r"\d+(?:[.,]\d{3})+", compact):
        return re.sub(r"\D", "", compact)
    return compact


def normalize_code(value: object) -> str:
    text = normalize_cell(value)
    if not text:
        return ""
    if re.fullmatch(r"[0-9][0-9.,\s]*", text):
        return re.sub(r"\D", "", text)
    return text


def sanitize_component(value: object) -> str:
    text = normalize_cell(value)
    text = text.replace("\\", "_").replace("/", "_")
    text = re.sub(r"\s+", " ", text).strip()
    return text


def build_source_candidates(
    source_base: Path,
    row: PlanillaRow,
    *,
    expediente_override: str | None = None,
) -> list[tuple[str, Path]]:
    tram = row.dig_tramite
    exp = expediente_override or row.dig_expediente
    anio = row.dig_anio
    id_tramite = row.dig_id_tramite

    candidates: list[tuple[str, Path]] = []
    if id_tramite:
        candidates.append(("structured_with_id_tramite", source_base / anio / exp / id_tramite / tram))
    candidates.append(("structured", source_base / anio / exp / tram))
    candidates.append(("direct_tramite", source_base / tram))
    candidates.append(("expediente_tramite", source_base / exp / tram))
    return candidates


def build_dest_path(
    dest_base: Path,
    row: PlanillaRow,
    include_id_tramite: bool,
    *,
    expediente_override: str | None = None,
) -> Path:
    expediente = expediente_override or row.dig_expediente
    parts = [dest_base, Path(row.dig_anio), Path(expediente)]
    if include_id_tramite and row.dig_id_tramite:
        parts.append(Path(row.dig_id_tramite))
    parts.append(Path(row.dig_tramite))
    result = parts[0]
    for part in parts[1:]:
        result = result / part
    return result


def count_files(folder: Path) -> int:
    total = 0
    try:
        for item in folder.rglob("*"):
            if item.is_file():
                total += 1
    except OSError:
        return total
    return total


def list_pdfs(folder: Path) -> list[Path]:
    pdfs: list[Path] = []
    try:
        for item in folder.rglob("*"):
            if item.is_file() and item.suffix.lower() == ".pdf":
                pdfs.append(item)
    except OSError:
        return pdfs
    return pdfs


def unique_dest_path(folder: Path, filename: str) -> Path:
    candidate = folder / filename
    if not candidate.exists():
        return candidate

    stem = Path(filename).stem
    suffix = Path(filename).suffix
    index = 2
    while True:
        alt = folder / f"{stem}_{index}{suffix}"
        if not alt.exists():
            return alt
        index += 1


class SourceFinder:
    def __init__(self, source_base: Path):
        self.source_base = source_base
        self._dir_index: dict[str, list[Path]] | None = None

    def _build_index(self) -> None:
        index: dict[str, list[Path]] = defaultdict(list)
        try:
            for path in self.source_base.rglob("*"):
                if path.is_dir():
                    index[path.name].append(path)
        except OSError:
            pass
        self._dir_index = dict(index)

    def find(self, row: PlanillaRow) -> tuple[Path | None, str, str]:
        for mode, candidate in build_source_candidates(self.source_base, row):
            if candidate.is_dir() and ensure_under_base(self.source_base, candidate):
                return candidate, mode, ""

        if self._dir_index is None:
            self._build_index()

        matches = []
        if self._dir_index is not None:
            matches = [
                path
                for path in self._dir_index.get(row.dig_tramite, [])
                if ensure_under_base(self.source_base, path)
            ]

        if len(matches) == 1:
            return matches[0], "recursive_unique", ""
        if len(matches) > 1:
            joined = "; ".join(str(path) for path in matches[:5])
            if len(matches) > 5:
                joined += f"; ... (+{len(matches) - 5})"
            return None, "ambiguous", joined

        return None, "missing", ""


def read_rows_from_excel(excel_path: Path, sheet_name: str | None = None) -> list[PlanillaRow]:
    workbook = load_workbook(excel_path, read_only=True, data_only=True)
    try:
        worksheet = workbook[sheet_name] if sheet_name else workbook[workbook.sheetnames[0]]

        header_row = None
        headers: list[object] = []
        for row_index, values in enumerate(worksheet.iter_rows(values_only=True), start=1):
            if any(value not in (None, "") for value in values):
                header_row = row_index
                headers = list(values)
                break

        if header_row is None:
            raise ValueError("El archivo Excel no contiene una fila de encabezados utilizable.")

        header_map: dict[str, int] = {}
        for idx, header in enumerate(headers):
            normalized = _normalize_header(header)
            if normalized and normalized not in header_map:
                header_map[normalized] = idx

        required = ["dig_id", "dig_anio", "dig_expediente", "dig_id_tramite", "dig_tramite"]
        missing = [name for name in required if name not in header_map]
        if missing:
            raise ValueError("Faltan columnas requeridas en Excel: " + ", ".join(missing))

        rows: list[PlanillaRow] = []
        for row_index, values in enumerate(
            worksheet.iter_rows(min_row=header_row + 1, values_only=True),
            start=header_row + 1,
        ):
            if values is None or all(value in (None, "") for value in values):
                continue

            def cell(name: str) -> object:
                idx = header_map[name]
                if idx >= len(values):
                    return None
                return values[idx]

            rows.append(
                PlanillaRow(
                    row_number=row_index,
                    dig_id=normalize_code(cell("dig_id")),
                    dig_anio=sanitize_component(cell("dig_anio")),
                    dig_expediente=sanitize_component(cell("dig_expediente")),
                    dig_id_tramite=sanitize_component(cell("dig_id_tramite")),
                    dig_tramite=sanitize_component(cell("dig_tramite")),
                )
            )
        return rows
    finally:
        workbook.close()


def filter_rows_by_dig_id_tramite(
    rows: list[PlanillaRow],
    dig_id_tramite: str | None,
) -> list[PlanillaRow]:
    target = normalize_code(dig_id_tramite or "")
    if not target:
        return rows
    return [row for row in rows if normalize_code(row.dig_id_tramite) == target]


def expediente_rank_key(expediente: str) -> tuple[int, str, str]:
    text = sanitize_component(expediente)
    match = re.fullmatch(r"^(.*?)(\d+)$", text)
    if match:
        prefix = match.group(1)
        number = int(match.group(2))
        return number, prefix, text
    return -1, text, text


def select_rows_with_highest_expediente(
    rows: list[PlanillaRow],
) -> tuple[list[PlanillaRow], list[PlanillaRow]]:
    grouped: dict[tuple[str, str], list[PlanillaRow]] = defaultdict(list)
    order: list[tuple[str, str]] = []
    for row in rows:
        key = (normalize_code(row.dig_id_tramite), normalize_code(row.dig_tramite))
        if key not in grouped:
            order.append(key)
        grouped[key].append(row)

    selected: list[PlanillaRow] = []
    discarded: list[PlanillaRow] = []
    for key in order:
        bucket = grouped[key]
        if len(bucket) == 1:
            selected.append(bucket[0])
            continue

        winner = max(
            bucket,
            key=lambda row: (
                expediente_rank_key(row.dig_expediente),
                -row.row_number,
            ),
        )
        selected.append(winner)
        discarded.extend(row for row in bucket if row is not winner)

    return selected, discarded


def build_destination_expediente_map(rows: list[PlanillaRow]) -> dict[str, str]:
    winners: dict[str, str] = {}
    for row in rows:
        key = normalize_code(row.dig_id_tramite)
        if not key:
            continue
        candidate = sanitize_component(row.dig_expediente)
        current = winners.get(key, "")
        if not current or expediente_rank_key(candidate) > expediente_rank_key(current):
            winners[key] = candidate
    return winners


def copy_planilla_tree(
    *,
    source_base: Path,
    dest_base: Path,
    row: PlanillaRow,
    include_id_tramite: bool,
    dry_run: bool,
    source_finder: SourceFinder,
    expediente_override: str | None = None,
) -> dict[str, object]:
    effective_expediente = expediente_override or row.dig_expediente
    dst = build_dest_path(
        dest_base,
        row,
        include_id_tramite=include_id_tramite,
        expediente_override=effective_expediente,
    )

    result: dict[str, object] = {
        "row_number": row.row_number,
        "dig_id": row.dig_id,
        "dig_anio": row.dig_anio,
        "dig_expediente": row.dig_expediente,
        "source_expediente": row.dig_expediente,
        "dest_expediente": effective_expediente,
        "dig_id_tramite": row.dig_id_tramite,
        "dig_tramite": row.dig_tramite,
        "source_path": "",
        "dest_path": str(dst),
        "source_match": "",
        "status": "",
        "files_found": 0,
        "pdfs_found": 0,
        "pdfs_copied": 0,
        "message": "",
    }

    required_fields = [row.dig_anio, row.dig_expediente, row.dig_tramite]
    if any(not field for field in required_fields):
        result["status"] = "INVALID_ROW"
        result["message"] = "faltan campos requeridos para construir la ruta"
        return result

    if include_id_tramite and not row.dig_id_tramite:
        result["message"] = "dig_id_tramite vacio, se omitio en la ruta"

    source_dir, source_match, detail = source_finder.find(row)
    result["source_match"] = source_match
    if detail:
        result["message"] = detail

    if source_match == "ambiguous":
        result["status"] = "AMBIGUOUS_SOURCE"
        return result
    if source_dir is None or not source_dir.exists() or not source_dir.is_dir():
        result["status"] = "MISSING_SOURCE"
        return result

    result["source_path"] = str(source_dir)
    result["files_found"] = count_files(source_dir)
    pdfs = list_pdfs(source_dir)
    result["pdfs_found"] = len(pdfs)

    if not pdfs:
        result["status"] = "WOULD_SKIP_NO_PDFS" if dry_run else "NO_PDFS"
        return result

    if dry_run:
        result["status"] = "WOULD_COPY"
        return result

    if not ensure_under_base(dest_base, dst):
        result["status"] = "INVALID_DESTINATION"
        result["message"] = "el destino calculado queda fuera de dest-base"
        return result

    dst.mkdir(parents=True, exist_ok=True)
    copied = 0
    for pdf in pdfs:
        target = unique_dest_path(dst, pdf.name)
        shutil.copy2(pdf, target)
        copied += 1

    result["pdfs_copied"] = copied
    result["status"] = "COPIED"
    return result


def write_manifest(path: Path, rows: list[dict[str, object]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=[
                "row_number",
                "dig_id",
                "dig_anio",
                "dig_expediente",
                "source_expediente",
                "dest_expediente",
                "dig_id_tramite",
                "dig_tramite",
                "source_path",
                "dest_path",
                "source_match",
                "status",
                "files_found",
                "pdfs_found",
                "pdfs_copied",
                "message",
            ],
        )
        writer.writeheader()
        for row in rows:
            writer.writerow(row)


def _apply_sheet_header_style(worksheet) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")


def write_excel_report(
    path: Path,
    summary: dict[str, object],
    detail_rows: list[dict[str, object]],
    *,
    include_id_tramite: bool,
) -> None:
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Resumen"

    summary_sheet.append(["Campo", "Valor"])
    for key, value in summary.items():
        summary_sheet.append([key, value])
    _apply_sheet_header_style(summary_sheet)
    summary_sheet.freeze_panes = "A2"
    summary_sheet.column_dimensions["A"].width = 28
    summary_sheet.column_dimensions["B"].width = 80

    detail_sheet = workbook.create_sheet("Detalle")
    detail_headers = [
        "row_number",
        "dig_id",
        "dig_anio",
        "dig_expediente",
        "source_expediente",
        "dest_expediente",
        "dig_id_tramite",
        "dig_tramite",
        "source_path",
        "dest_path",
        "source_match",
        "status",
        "files_found",
        "pdfs_found",
        "pdfs_copied",
        "message",
    ]
    detail_sheet.append(detail_headers)
    for row in detail_rows:
        detail_sheet.append([row.get(header, "") for header in detail_headers])
    _apply_sheet_header_style(detail_sheet)
    detail_sheet.freeze_panes = "A2"
    for column, width in {
        "A": 12,
        "B": 14,
        "C": 12,
        "D": 18,
        "E": 18,
        "F": 18,
        "G": 18,
        "H": 18,
        "I": 50,
        "J": 50,
        "K": 18,
        "L": 18,
        "M": 12,
        "N": 12,
        "O": 12,
        "P": 40,
    }.items():
        detail_sheet.column_dimensions[column].width = width

    grouped: dict[tuple[str, str, str], dict[str, object]] = {}
    for row in detail_rows:
        expediente = str(row.get("dest_expediente", "") or row.get("dig_expediente", ""))
        key = (
            str(row.get("dig_anio", "")),
            expediente,
            str(row.get("dig_id_tramite", "")) if include_id_tramite else "",
        )
        bucket = grouped.setdefault(
            key,
            {
                "dig_anio": key[0],
                "dig_expediente": key[1],
                "dig_id_tramite": key[2],
                "tramites": 0,
                "pdfs_found": 0,
                "pdfs_copied": 0,
            },
        )
        bucket["tramites"] = int(bucket["tramites"]) + 1
        bucket["pdfs_found"] = int(bucket["pdfs_found"]) + int(row.get("pdfs_found", 0) or 0)
        bucket["pdfs_copied"] = int(bucket["pdfs_copied"]) + int(row.get("pdfs_copied", 0) or 0)

    grouped_sheet = workbook.create_sheet("Agrupado")
    grouped_headers = ["dig_anio", "dig_expediente", "dig_id_tramite", "tramites", "pdfs_found", "pdfs_copied"]
    grouped_sheet.append(grouped_headers)
    for _, bucket in sorted(
        grouped.items(),
        key=lambda item: (-int(item[1]["tramites"]), item[0][0], item[0][1], item[0][2]),
    ):
        grouped_sheet.append([bucket.get(h, "") for h in grouped_headers])
    _apply_sheet_header_style(grouped_sheet)
    grouped_sheet.freeze_panes = "A2"
    for column, width in {
        "A": 12,
        "B": 18,
        "C": 18,
        "D": 12,
        "E": 12,
        "F": 12,
    }.items():
        grouped_sheet.column_dimensions[column].width = width

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def summarize_results(detail_rows: list[dict[str, object]]) -> dict[str, int]:
    summary = {
        "total": 0,
        "copied": 0,
        "would_copy": 0,
        "missing_source": 0,
        "no_pdfs": 0,
        "invalid_row": 0,
        "ambiguous_source": 0,
        "invalid_destination": 0,
    }

    for row in detail_rows:
        summary["total"] += 1
        status = str(row.get("status", ""))
        if status == "COPIED":
            summary["copied"] += 1
        elif status == "WOULD_COPY":
            summary["would_copy"] += 1
        elif status in {"MISSING_SOURCE", "WOULD_MISSING_SOURCE"}:
            summary["missing_source"] += 1
        elif status in {"NO_PDFS", "WOULD_SKIP_NO_PDFS"}:
            summary["no_pdfs"] += 1
        elif status == "INVALID_ROW":
            summary["invalid_row"] += 1
        elif status == "AMBIGUOUS_SOURCE":
            summary["ambiguous_source"] += 1
        elif status == "INVALID_DESTINATION":
            summary["invalid_destination"] += 1

    return summary


def format_timestamp() -> str:
    return datetime.now().isoformat(timespec="seconds")
